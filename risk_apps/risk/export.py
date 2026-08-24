from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from collections import defaultdict


HEADERS = [
    "Risk ID",
    "Risk Category",
    "Risk Type",
    "Risk Description",
    "Likelihood (1-5)",
    "Impact (1-5)",
    "Risk Score",
    "Risk Level",
    "Risk Owner",
    "Mitigation Plan",
    "Status",
    "Date Identified",
    "Next Review Date",
    "Business Unit",
    "Fraud Related",
    "ESG Area",
    "Continuity Dependency",
]

START_ROW = 5
START_COL = 2  # Column B


def format_description(risk):
    return (
        f"Event: {risk.event}\n\n"
        f"Cause: {risk.cause}\n\n"
        f"Impact: {getattr(risk, 'impact_description', '')}"
    )


def write_sheet(ws, risks, title=None):
    """
    Reusable sheet writer (Enterprise standard)
    """

    # Optional title (Row 2)
    if title:
        ws.cell(row=2, column=START_COL, value=title).font = Font(bold=True, size=14)

    # Write headers
    for col_index, header in enumerate(HEADERS, start=START_COL):
        cell = ws.cell(row=START_ROW, column=col_index, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data
    row_num = START_ROW + 1

    for risk in risks:
        ws.cell(row=row_num, column=2, value=risk.risk_id)
        ws.cell(row=row_num, column=3, value=risk.category.name if risk.category else "")
        ws.cell(row=row_num, column=4, value=risk.risk_type)
        ws.cell(row=row_num, column=5, value=format_description(risk))
        ws.cell(row=row_num, column=6, value=risk.likelihood.rating if risk.likelihood else "")
        ws.cell(row=row_num, column=7, value=risk.impact.rating if risk.impact else "")
        ws.cell(row=row_num, column=8, value=risk.risk_score)
        ws.cell(row=row_num, column=9, value=risk.risk_level)
        ws.cell(row=row_num, column=10, value=risk.risk_owner)
        ws.cell(row=row_num, column=11, value=risk.mitigation_plan)
        ws.cell(row=row_num, column=12, value=risk.status)
        ws.cell(row=row_num, column=13, value=risk.date_identified)
        ws.cell(row=row_num, column=14, value=risk.next_review_date)
        ws.cell(row=row_num, column=15, value=risk.business_unit)
        ws.cell(row=row_num, column=16, value="Yes" if risk.is_fraud_related else "No")
        ws.cell(row=row_num, column=17, value=risk.esg_area)
        ws.cell(row=row_num, column=18, value=risk.continuity_dependency)

        row_num += 1

    # Auto column width (optimized)
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max_length + 2, 45)


def build_risk_register_workbook(queryset):
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # =========================
    # 1️⃣ MASTER SHEET (ALL RISKS)
    # =========================
    ws_all = wb.create_sheet(title="Risk Register")
    write_sheet(ws_all, queryset, title="Enterprise Risk Register")

    # =========================
    # 2️⃣ GROUP BY CATEGORY
    # =========================
    category_map = defaultdict(list)

    for risk in queryset:
        key = risk.category.name if risk.category else "Uncategorized"
        category_map[key].append(risk)

    # =========================
    # 3️⃣ CREATE CATEGORY SHEETS
    # =========================
    for category_name, risks in category_map.items():

        # Excel sheet name limit = 31 chars
        sheet_name = category_name[:31]

        ws = wb.create_sheet(title=sheet_name)

        write_sheet(
            ws,
            risks,
            title=f"{category_name} Risk Register"
        )

    return wb
