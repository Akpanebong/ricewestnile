from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


HEADERS = [
    "Risk ID", "Category", "Type", "Description",
    "Likelihood", "Impact", "Score", "Level",
    "Owner", "Mitigation", "Status", "Date"
]


def build_workbook(queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Risk Register"

    # headers
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # data
    for row, risk in enumerate(queryset, 2):
        ws.cell(row=row, column=1, value=risk.risk_id)
        ws.cell(row=row, column=2, value=risk.category.name if risk.category else "")
        ws.cell(row=row, column=3, value=risk.risk_type)
        ws.cell(row=row, column=4, value=risk.event)
        ws.cell(row=row, column=5, value=risk.likelihood.rating if risk.likelihood else "")
        ws.cell(row=row, column=6, value=risk.impact.rating if risk.impact else "")
        ws.cell(row=row, column=7, value=risk.risk_score)
        ws.cell(row=row, column=8, value=risk.risk_level)

    return wb