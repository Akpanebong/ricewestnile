
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from core.project_models import ProjectBudget, Project
from procurement.procureapp.models import ProcurementPlan, ProcurementPlanItem
from core.services import display_amount_from_ugx, get_user_currency
from procurement.procureapp.utils import render_pdf


@login_required
def master_procurement_plan(request):
    """Master plan filtered by project, fiscal year, budget period, status, and delivery date."""
    filters = {
        'year': request.GET.get('year') or str(timezone.now().year),
        'project': request.GET.get('project') or '',
        'period': request.GET.get('period') or '',
        'status': request.GET.get('status') or '',
        'delivery_from': request.GET.get('delivery_from') or '',
        'delivery_to': request.GET.get('delivery_to') or '',
    }

    plans_qs = ProcurementPlan.objects.select_related('project', 'project__unit', 'project__unit__department')
    if filters['year']:
        plans_qs = plans_qs.filter(fiscal_year=filters['year'])
    if filters['project']:
        plans_qs = plans_qs.filter(project_id=filters['project'])
    if filters['period']:
        plans_qs = plans_qs.filter(budget_period=filters['period'])
    if filters['status']:
        plans_qs = plans_qs.filter(status=filters['status'])
    else:
        plans_qs = plans_qs.exclude(status__iexact='Draft')

    items_qs = (
        ProcurementPlanItem.objects
        .select_related(
            'product',
            'procurement',
            'procurement__project',
            'procurement__project__unit',
            'procurement__project__unit__department',
        )
        .filter(procurement__in=plans_qs)
    )
    if filters['delivery_from']:
        items_qs = items_qs.filter(delivery_date__gte=filters['delivery_from'])
    if filters['delivery_to']:
        items_qs = items_qs.filter(delivery_date__lte=filters['delivery_to'])

    items = list(items_qs.order_by(
        'procurement__project__name',
        'procurement__fiscal_year',
        'procurement__budget_period',
        'delivery_date',
        'description',
    ))
    grand_total = sum((item.qty or 0) * (item.est_unit_cost or 0) for item in items)

    export_params = request.GET.copy()
    export_params.pop('format', None)
    export_querystring = export_params.urlencode()

    available_years = (
        ProcurementPlan.objects
        .exclude(fiscal_year__isnull=True)
        .exclude(fiscal_year='')
        .values_list('fiscal_year', flat=True)
        .distinct()
        .order_by('-fiscal_year')
    )

    context = {
        'year': filters['year'],
        'filters': filters,
        'items': items,
        'grand_total': grand_total,
        'projects': Project.objects.order_by('name'),
        'available_years': available_years,
        'period_choices': ProjectBudget.PERIOD_CHOICES,
        'status_choices': ProcurementPlan._meta.get_field('status').choices,
        'export_querystring': export_querystring,
        'display_currency_code': get_user_currency(request),
        'org_name': 'RICE WEST NILE',
        'request': request,
    }

    export = request.GET.get('format')
    if export == 'pdf':
        return render_pdf('procurement/master_plan_pdf.html', context, f'Master_Procurement_Plan_{filters["year"]}.pdf')
    if export == 'excel':
        return _export_master_plan_excel(context, request)

    return render(request, 'procurement/master_plan.html', context)


def _export_master_plan_excel(ctx, request):
    """Excel export helper using openpyxl."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Master Plan {ctx['year']}"

    row = 1
    ws.cell(row=row, column=1, value=f"{ctx['org_name']} - Master Procurement Plan {ctx['year']}")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
    row += 2

    filter_summary = [
        f"Project: {next((str(project) for project in ctx['projects'] if str(project.pk) == ctx['filters']['project']), 'All projects')}",
        f"Year: {ctx['filters']['year'] or 'All years'}",
        f"Period: {ctx['filters']['period'] or 'All periods'}",
        f"Status: {ctx['filters']['status'] or 'All statuses'}",
        f"Currency: {ctx['display_currency_code']}",
    ]
    ws.cell(row=row, column=1, value=' | '.join(filter_summary))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
    row += 2

    header = [
        "#",
        "Project",
        "Department",
        "Plan No.",
        "Fiscal Year",
        "Period",
        "Status",
        "Description",
        "Unit Measure",
        "Qty",
        f"Est. Unit Cost ({ctx['display_currency_code']})",
        f"Line Total ({ctx['display_currency_code']})",
        "Delivery Date",
    ]
    for col, h in enumerate(header, start=1):
        ws.cell(row=row, column=col, value=h)
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    row += 1

    for index, it in enumerate(ctx['items'], start=1):
        plan = it.procurement
        project = plan.project
        line_total = (it.qty or 0) * (it.est_unit_cost or 0)
        ws.cell(row=row, column=1, value=index)
        ws.cell(row=row, column=2, value=str(project) if project else 'Unassigned')
        ws.cell(row=row, column=3, value=str(project.department) if project and project.department else '')
        ws.cell(row=row, column=4, value=plan.number)
        ws.cell(row=row, column=5, value=plan.fiscal_year)
        ws.cell(row=row, column=6, value=plan.get_budget_period_display())
        ws.cell(row=row, column=7, value=plan.status)
        ws.cell(row=row, column=8, value=it.product.name if it.product_id else it.description)
        ws.cell(row=row, column=9, value=it.unit_measure)
        ws.cell(row=row, column=10, value=float(it.qty or 0))
        ws.cell(row=row, column=11, value=float(display_amount_from_ugx(it.est_unit_cost or 0, request)))
        ws.cell(row=row, column=12, value=float(display_amount_from_ugx(line_total, request)))
        ws.cell(row=row, column=13, value=it.delivery_date)
        row += 1

    ws.cell(row=row, column=11, value="GRAND TOTAL")
    ws.cell(row=row, column=11).font = Font(bold=True)
    ws.cell(row=row, column=12, value=float(display_amount_from_ugx(ctx['grand_total'], request)))
    ws.cell(row=row, column=12).font = Font(bold=True)

    # autosize columns
    for col in range(1, 14):
        max_len = 0
        for r in range(1, row + 1):
            v = ws.cell(r, col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 50)

    # return response
    from io import BytesIO
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    resp = HttpResponse(
        stream.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename=Master_Procurement_Plan_{ctx["year"]}.xlsx'
    return resp
