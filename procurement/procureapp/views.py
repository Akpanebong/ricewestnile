from django.contrib.auth.models import Group
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy, reverse
from django.views.generic import DetailView, ListView
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Count, DecimalField, F, Q, Sum, Value
from django.db.models.functions import Coalesce
from django.db import transaction
from django.core.exceptions import ValidationError
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from datetime import datetime, timedelta
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from django.core.mail import send_mail
from django.conf import settings
from account.models import Department, Profile, Unit
from core.project_models import ProjectBudget, Project
from core.models import CurrencyRate
from core.services import display_amount_from_ugx, get_user_currency, user_amount_to_ugx
from notification.models import NotificationRecipient
from notification.utils import notify
from finance_app.finance.models import AccountingForm, AdminExpenseNote, CashRequisition, CashRequisitionItem

from account.templatetags.custom_tags import has_group
from .models import (
    AuditLog, Supplier, ProcurementPlan, RFQ,
    Requisition, Product, RFQSendLog, PurchaseOrder, SupplierSpendReport)
from .forms import (
    ProcurementPlanForm,
    ProcurementPlanItemFormSet,
    ProductForm,
    SupplierRegistrationForm,
    SupplierSpendReportUpdateForm,
)
from .utils import render_pdf, head_of_procurement_required, send_html_email


# Helper - flexible date parsing
def parse_date_safe(date_str):
    if not date_str:
        return None
    for fmt in ("%Y-%m-%d", "%b. %d, %Y", "%B %d, %Y", "%b %d, %Y"):
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None


def normalize_date(value):
    if isinstance(value, datetime):
        return value.date()
    return value


def get_formset_total(formset):
    total = 0
    for item_form in formset:
        if not item_form.cleaned_data or item_form.cleaned_data.get('DELETE'):
            continue
        qty = item_form.cleaned_data.get('qty') or 0
        est_unit_cost = item_form.cleaned_data.get('est_unit_cost') or 0
        total += qty * est_unit_cost
    return total


def add_validation_messages(request, error):
    if hasattr(error, 'messages'):
        for message in error.messages:
            messages.error(request, message)
    else:
        messages.error(request, str(error))


def get_project_budget_data():
    current_year = ProjectBudget.current_fiscal_year()
    data = {}

    for project in Project.objects.prefetch_related('budgets').order_by('name'):
        budgets = list(project.budgets.all())
        fiscal_years = sorted({budget.fiscal_year for budget in budgets}, reverse=True)
        default_year = current_year if current_year in fiscal_years else (fiscal_years[0] if fiscal_years else current_year)

        data[str(project.pk)] = {
            'donor': project.donor,
            'fiscal_year': default_year,
            'budgets': [
                {
                    'fiscal_year': budget.fiscal_year,
                    'period': budget.period,
                    'amount': str(budget.budget_amount),
                }
                for budget in budgets
            ],
        }

    return data


@login_required(login_url='login')
def dashboard(request):
    if not (has_group(request.user, 'Procurement') or has_group(request.user, 'ED')  or request.user.is_superuser):
        return redirect(reverse_lazy("profile_update", kwargs={'pk': request.user.pk, 'slug': request.user.slug}))

    money_field = DecimalField(max_digits=18, decimal_places=2)
    zero_money = Value(0, output_field=money_field)

    procurement_total = ProcurementPlan.objects.aggregate(
        total=Coalesce(Sum('total_amount'), zero_money)
    )['total']
    requisition_total = Requisition.objects.aggregate(
        total=Coalesce(Sum('total'), zero_money),
        awarded=Coalesce(Sum('awarded_amount'), zero_money),
    )
    purchase_order_total = PurchaseOrder.objects.aggregate(
        total=Coalesce(Sum('final_amount'), zero_money)
    )['total']
    supplier_spend_total = SupplierSpendReport.objects.aggregate(
        total=Coalesce(Sum('total_spent'), zero_money)
    )['total']
    project_budget_total = ProjectBudget.objects.aggregate(
        total=Coalesce(Sum('budget_amount'), zero_money)
    )['total']
    cash_requisition_total = CashRequisitionItem.objects.aggregate(
        total=Coalesce(Sum(F('quantity') * F('unit_cost'), output_field=money_field), zero_money)
    )['total']
    admin_expense_total = AdminExpenseNote.objects.aggregate(
        total=Coalesce(Sum('proposed_budget'), zero_money)
    )['total']

    procurement_status = list(
        ProcurementPlan.objects.values('status').annotate(count=Count('id')).order_by('status')
    )
    requisition_status = list(
        Requisition.objects.values('status').annotate(count=Count('id')).order_by('status')
    )
    rfq_status = list(
        RFQ.objects.values('status').annotate(count=Count('id')).order_by('status')
    )
    finance_status = list(
        CashRequisition.objects.values('status').annotate(count=Count('id')).order_by('status')
    )

    department_mix = list(
        ProcurementPlan.objects
        .filter(project__unit__department__isnull=False)
        .values(name=F('project__unit__department__name'))
        .annotate(count=Count('id'), total=Coalesce(Sum('total_amount'), zero_money))
        .order_by('-total')[:6]
    )
    budget_utilization = list(
        ProjectBudget.objects
        .select_related('project')
        .order_by('-fiscal_year', 'project__name')[:6]
    )
    top_suppliers = list(
        Supplier.objects
        .annotate(total_spent=Coalesce(Sum('purchase_orders__final_amount', filter=Q(purchase_orders__sent=True)), zero_money))
        .filter(total_spent__gt=0)
        .order_by('-total_spent')[:6]
    )
    recent_purchase_orders = (
        PurchaseOrder.objects
        .select_related('supplier', 'requisition', 'procurement_plan')
        .order_by('-created_at')[:5]
    )
    pending_work = [
        {
            'label': 'Procurement plans awaiting review or approval',
            'count': ProcurementPlan.objects.exclude(status='Approved').count(),
            'url': reverse('requisition_list'),
            'icon': 'fa-clipboard-check',
            'tone': 'primary',
        },
        {
            'label': 'Requisitions not fully approved',
            'count': Requisition.objects.exclude(status='Approved').count(),
            'url': reverse('req_list'),
            'icon': 'fa-file-signature',
            'tone': 'success',
        },
        {
            'label': 'RFQs pending supplier response',
            'count': RFQ.objects.exclude(status='Responded').count(),
            'url': reverse('rfq_list'),
            'icon': 'fa-envelope-open-text',
            'tone': 'warning',
        },
        {
            'label': 'Purchase orders not sent',
            'count': PurchaseOrder.objects.filter(sent=False).count(),
            'url': reverse('purchase_order'),
            'icon': 'fa-truck-fast',
            'tone': 'danger',
        },
    ]

    recent_activities = []
    for r in ProcurementPlan.objects.select_related('project').order_by('-created_at')[:4]:
        recent_activities.append({
            'type': 'ProcurementPlan',
            'ref': r.number,
            'source': r.project or r.department or '-',
            'date': r.created_at,
            'status': r.status
        })
    for po in Requisition.objects.select_related('procurement').order_by('-created_at')[:4]:
        recent_activities.append({
            'type': 'Requisition',
            'ref': po.number,
            'source': po.procurement if po.procurement else '-',
            'date': po.created_at,
            'status': po.status
        })
    for inv in RFQ.objects.select_related('req').order_by('-created_at')[:4]:
        recent_activities.append({
            'type': 'RFQ',
            'ref': inv.reference_no,
            'source': inv.req if inv.req else '-',
            'date': inv.created_at,
            'status': inv.status
        })
    for order in PurchaseOrder.objects.select_related('supplier').order_by('-created_at')[:4]:
        recent_activities.append({
            'type': 'Purchase Order',
            'ref': order.po_number or order.pk,
            'source': order.supplier,
            'date': order.created_at,
            'status': 'Sent' if order.sent else 'Draft',
        })

    recent_activities = sorted(recent_activities, key=lambda x: x['date'], reverse=True)[:10]

    model_coverage = [
        {'group': 'Procurement', 'items': [
            {'name': 'Products', 'count': Product.objects.count(), 'url': reverse('product_list')},
            {'name': 'Procurement Plans', 'count': ProcurementPlan.objects.count(), 'url': reverse('requisition_list')},
            {'name': 'Plan Items', 'count': ProcurementPlan.objects.aggregate(count=Count('items'))['count'], 'url': reverse('requisition_list')},
            {'name': 'Requisitions', 'count': Requisition.objects.count(), 'url': reverse('req_list')},
            {'name': 'RFQs', 'count': RFQ.objects.count(), 'url': reverse('rfq_list')},
            {'name': 'RFQ Send Logs', 'count': RFQSendLog.objects.count(), 'url': reverse('rfq_list')},
            {'name': 'Purchase Orders', 'count': PurchaseOrder.objects.count(), 'url': reverse('purchase_order')},
            {'name': 'Supplier Spend Reports', 'count': SupplierSpendReport.objects.filter(payment_status='Paid').count(), 'url': reverse('supplier_report')},
        ]},
        {'group': 'Suppliers & Organization', 'items': [
            {'name': 'Suppliers', 'count': Supplier.objects.count(), 'url': reverse('supplier_list')},
            {'name': 'Active Suppliers', 'count': Supplier.objects.filter(active=True).count(), 'url': reverse('supplier_list')},
            {'name': 'Departments', 'count': Department.objects.count(), 'url': reverse('department_list')},
            {'name': 'Units', 'count': Unit.objects.count(), 'url': reverse('department_list')},
            {'name': 'Projects', 'count': Project.objects.count(), 'url': reverse('procurement_core:project_list')},
            {'name': 'Project Budgets', 'count': ProjectBudget.objects.count(), 'url': reverse('procurement_core:project_list')},
            {'name': 'Users', 'count': Profile.objects.count(), 'url': reverse('profile_list')},
        ]},
        {'group': 'Finance, Controls & System', 'items': [
            {'name': 'Cash Requisitions', 'count': CashRequisition.objects.count(), 'url': reverse('finance:list_cash_req')},
            {'name': 'Admin Expense Notes', 'count': AdminExpenseNote.objects.count(), 'url': reverse('finance:dashboard')},
            {'name': 'Accounting Forms', 'count': AccountingForm.objects.count(), 'url': reverse('finance:dashboard')},
            {'name': 'Notifications', 'count': NotificationRecipient.objects.count(), 'url': reverse('notifications:list')},
            {'name': 'Unread Notifications', 'count': NotificationRecipient.objects.filter(is_read=False, is_deleted=False).count(), 'url': reverse('notifications:list')},
            {'name': 'Audit Logs', 'count': AuditLog.objects.count(), 'url': reverse('notifications:list')},
            {'name': 'Currency Rates', 'count': CurrencyRate.objects.count(), 'url': reverse('core:currency_settings')},
        ]},
    ]

    return render(request, 'procurement/dashboard.html', {
        'pr_pending': ProcurementPlan.objects.count(),
        'po_count': Requisition.objects.count(),
        'supplier_count': Supplier.objects.count(),
        'po_pending': RFQ.objects.count(),
        'kpis': [
            {'label': 'Procurement Pipeline', 'value': ProcurementPlan.objects.count(), 'amount': procurement_total, 'icon': 'fa-diagram-project', 'tone': 'primary', 'url': reverse('requisition_list')},
            {'label': 'Requisition Value', 'value': Requisition.objects.count(), 'amount': requisition_total['total'], 'icon': 'fa-file-signature', 'tone': 'success', 'url': reverse('req_list')},
            {'label': 'Purchase Orders', 'value': PurchaseOrder.objects.count(), 'amount': purchase_order_total, 'icon': 'fa-cart-shopping', 'tone': 'info', 'url': reverse('purchase_order')},
            {'label': 'Supplier Spend', 'value': SupplierSpendReport.objects.filter(payment_status='Paid').values('supplier').distinct().count(), 'amount': supplier_spend_total, 'icon': 'fa-handshake', 'tone': 'warning', 'url': reverse('supplier_report')},
            {'label': 'Project Budgets', 'value': ProjectBudget.objects.count(), 'amount': project_budget_total, 'icon': 'fa-scale-balanced', 'tone': 'dark', 'url': reverse('procurement_core:project_list')},
            {'label': 'Finance Requests', 'value': CashRequisition.objects.count(), 'amount': cash_requisition_total, 'icon': 'fa-money-bill-transfer', 'tone': 'danger', 'url': reverse('finance:list_cash_req')},
        ],
        'status_sections': [
            {'title': 'Procurement Plans', 'rows': procurement_status},
            {'title': 'Requisitions', 'rows': requisition_status},
            {'title': 'RFQs', 'rows': rfq_status},
            {'title': 'Cash Requisitions', 'rows': finance_status},
        ],
        'chart_labels': [item['status'] or 'Unspecified' for item in procurement_status],
        'chart_values': [item['count'] for item in procurement_status],
        'rfq_chart_labels': [item['status'] or 'Unspecified' for item in rfq_status],
        'rfq_chart_values': [item['count'] for item in rfq_status],
        'pending_work': pending_work,
        'department_mix': department_mix,
        'budget_utilization': budget_utilization,
        'top_suppliers': top_suppliers,
        'recent_purchase_orders': recent_purchase_orders,
        'model_coverage': model_coverage,
        'finance_summary': {
            'cash_requisition_total': cash_requisition_total,
            'admin_expense_total': admin_expense_total,
            'accounting_count': AccountingForm.objects.count(),
            'approved_cash_requisitions': CashRequisition.objects.filter(status='approved').count(),
        },
        'recent_activities': recent_activities,
    })


@login_required(login_url='login')
def supplier_list(request):
    suppliers = Supplier.objects.all().order_by('full_name', 'organisation_name', 'title')
    return render(request, 'procurement/supplier_list.html', {
        'object_list': suppliers,
        'active_count': suppliers.filter(active=True).count(),
    })


@login_required(login_url='login')
def supplier_register(request):
    form = SupplierRegistrationForm(request.POST or None, request.FILES or None)
    if form.is_valid():
        supplier = form.save()
        messages.success(request, 'Supplier registered successfully.')
        return redirect('update_supplier', pk=supplier.pk, phone=supplier.phone or 'none')

    return render(request, 'procurement/supplier_register.html', {
        'form': form,
        'is_update': False,
    })


@login_required(login_url='login')
def update_supplier(request, pk, phone=None):
    supplier = get_object_or_404(Supplier, pk=pk)
    form = SupplierRegistrationForm(request.POST or None, request.FILES or None, instance=supplier)
    if form.is_valid():
        form.save()
        messages.success(request, 'Supplier updated successfully.')
        return redirect('supplier_list')

    return render(request, 'procurement/supplier_register.html', {
        'form': form,
        'supplier': supplier,
        'is_update': True,
    })


@login_required(login_url='login')
def activate_supplier(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    supplier.active = not supplier.active
    supplier.save(update_fields=['active'])
    state = 'activated' if supplier.active else 'deactivated'
    messages.success(request, f'Supplier {state} successfully.')
    return redirect('update_supplier', pk=supplier.pk, phone=supplier.phone or 'none')


@login_required(login_url='login')
def trash_supplier(request, pk, phone=None, title=None):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        supplier.delete()
        messages.success(request, 'Supplier deleted successfully.')
        return redirect('supplier_list')

    return render(request, 'delete_confirmation.html', {
        'delete': supplier,
        'cancel_url': reverse('supplier_list'),
    })


@login_required(login_url='login')
def get_product_cost(request):
    """AJAX endpoint to return product estimated unit cost."""
    product_id = request.GET.get('product_id')
    try:
        product = Product.objects.get(id=product_id)
        return JsonResponse({'est_unit_cost': float(product.est_unit_cost)})
    except Product.DoesNotExist:
        return JsonResponse({'error': 'Product not found'}, status=404)


def notify_programs_head(request, procurement):
    try:
        programs_department = Department.objects.get(name__iexact='Programs')
        programs_head = programs_department.head

        if not programs_head:
            return False
        # In-app notification
        action_url = request.build_absolute_uri(
            reverse("requisition_detail", args=[procurement.pk]))

        # Create in-app notification
        reject_notify = f"""
                        The procurement plan with {procurement.number} submitted on {procurement.created_at.date()} was rejected by the Executive Director.
                        Please review and revert. Note that the requester is also notified.
                        \n THANK YOU.
                        """
        notify(
            request=request,
            users=programs_head,
            title="Procurement Plan Review",
            message=(
                f"Procurement Plan {procurement.number} "
                f"has been submitted and awaits your review."
            ) if procurement.status == 'Pending'
            else reject_notify,
            category="info",
            source_app="Procurement",
            action_url=action_url,
        )

        # Email notification
        if programs_head.email:
            send_html_email(
                request=request,
                subject="Procurement Plan Review Required",
                recipient=programs_head.email,
                title=f"Procurement Plan ({procurement.number}) Requires Review",
                message=(
                    f"A procurement plan has been submitted by "
                    f"{request.user.get_full_name() or request.user.username} "
                    f"and requires your review."
                ),
                relative_link=action_url,
            )

        return True

    except Department.DoesNotExist:
        messages.warning(
            request,
            "Programs Department has not been configured."
        )
        return False

    except Exception as e:
        print(f"Programs notification error: {e}")
        return False


@login_required(login_url='login')
def procurement_create(request):
    """Create ProcurementPlan and its items."""
    if not request.user.is_superuser and not Project.objects.filter(project_head=request.user).exists():
        messages.warning(request, "Only Project Heads can create Procurement Plans.")
        return redirect('dashboard')

    if request.method == 'POST':
        form = ProcurementPlanForm(request.POST or None, user=request.user)
        formset = ProcurementPlanItemFormSet(request.POST, prefix='items')

        if form.is_valid() and formset.is_valid():
            procurement = form.save(commit=False)
            procurement.requester = procurement.created_by = request.user
            proposed_total = get_formset_total(formset)

            try:
                project_budget = procurement.validate_budget_limit(proposed_total)
            except ValidationError as error:
                add_validation_messages(request, error)
            else:
                with transaction.atomic():
                    procurement.donor = procurement.project.donor if procurement.project else ''
                    procurement.project_budget = project_budget
                    procurement.budget = project_budget.budget_amount
                    procurement.save()

                    formset.instance = procurement
                    formset.save()

                    procurement.update_total()

                    notify_programs_head(request, procurement)
                    print(notify_programs_head(request, procurement))

                messages.success(request, "Procurement Plan created successfully with all items.")
                return redirect('requisition_list')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = ProcurementPlanForm(user=request.user)
        formset = ProcurementPlanItemFormSet(prefix='items')

    return render(request, 'procurement/procurement_form.html', {
        'form': form,
        'formset': formset,
        'project_budget_data': get_project_budget_data(),
    })


@login_required(login_url='login')
def procurement_update(request, pk):
    plan = get_object_or_404(ProcurementPlan, pk=pk)


    if plan.is_locked and not request.user.is_superuser:
        messages.warning(request, "Approved procurement plans can only be edited by an admin.")
        return redirect('requisition_detail', pk=plan.pk)

    if request.method == 'POST':
        form = ProcurementPlanForm(request.POST, instance=plan, user=request.user)
        formset = ProcurementPlanItemFormSet(request.POST, prefix='items', instance=plan)

        if form.is_valid() and formset.is_valid():
            procurement = form.save(commit=False)
            procurement.updated_by = request.user
            proposed_total = get_formset_total(formset)

            # in case plan is rejected and lated updated, also update status
            if procurement.status == 'Rejected':
                procurement.status = 'Pending'

            try:
                project_budget = procurement.validate_budget_limit(proposed_total)
            except ValidationError as error:
                add_validation_messages(request, error)
            else:
                with transaction.atomic():
                    procurement.donor = procurement.project.donor if procurement.project else ''
                    procurement.project_budget = project_budget
                    procurement.budget = project_budget.budget_amount
                    procurement.save()

                    formset.instance = procurement
                    formset.save()

                    procurement.update_total()

                messages.success(request, "ProcurementPlan updated successfully with all items.")
                return redirect('requisition_list')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = ProcurementPlanForm(instance=plan, user=request.user)
        formset = ProcurementPlanItemFormSet(prefix='items', instance=plan)

    return render(request, 'procurement/procurement_form.html', {
        'form': form,
        'formset': formset,
        'is_update': True,
        'project_budget_data': get_project_budget_data(),
    })


class ProcurementPlanListView(ListView):
    model = ProcurementPlan
    template_name = "procurement/procurement_list.html"
    context_object_name = "object_list"
    paginate_by = 10
    ordering = ['-created_at']

    def get_queryset(self):
        queryset = (
            ProcurementPlan.objects
            .select_related(
                # 'department',
                'project',
                'requester'
            )
            .prefetch_related('items')
            .order_by('-created_at')
        )

        search = self.request.GET.get('search')
        status = self.request.GET.get('status')

        if search:
            queryset = queryset.filter(
                Q(number__icontains=search) |
                Q(project__name__icontains=search) |
                Q(department__name__icontains=search) |
                Q(activity_name__icontains=search) |
                Q(donor__icontains=search) |
                Q(fiscal_year__icontains=search)
            )

        if status:
            queryset = queryset.filter(status=status)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        queryset = self.get_queryset()

        context['approved_count'] = queryset.filter(
            status='Approved'
        ).count()

        context['pending_count'] = queryset.filter(
            status='Pending'
        ).count()

        context['rejected_count'] = queryset.filter(
            status='Rejected'
        ).count()

        context['grand_total'] = (
            queryset.aggregate(
                total=Sum('total_amount')
            )['total'] or 0
        )

        context['status_filter'] = self.request.GET.get(
            'status', ''
        )

        context['search_query'] = self.request.GET.get(
            'search', ''
        )

        return context


class ProcurementPlanDetailView(DetailView):
    model = ProcurementPlan
    template_name = "procurement/procurement_detail.html"

    def get(self, request, *args, **kwargs):
        procurement = self.get_object()

        approvals = 0

        if procurement.reviewed_by:
            approvals += 1

        if procurement.approved_by:
            approvals += 1

        total_stages = 2
        progress_percent = int((approvals / total_stages) * 100)
        context = {
            "procurement": procurement,
            "items": procurement.items.select_related(),
            "approvals": approvals,
            "total_stages": total_stages,
            "progress_percent": progress_percent,
            'ed': Profile.objects.get(groups__name__iexact='ED'),
            "programs_head": get_object_or_404(Department, name="Programs").head if get_object_or_404(Department, name="Programs").head else None,
            "organization_name": "RICE West Nile",
            "total_amount": procurement.total_amount,
        }

        if request.GET.get("format") == "pdf":
            return render_pdf("procurement/procurement_pdf.html", context,
                              f"ProcurementPlan_{procurement.number}.pdf", request=request,)

        return self.render_to_response(context)

def trash_procurement(request, pk, slug):
    procurement = get_object_or_404(ProcurementPlan, pk=pk, slug=slug)
    if procurement.is_locked and not request.user.is_superuser:
        messages.warning(request, "Approved procurement plans can only be deleted by an admin.")
        return redirect('requisition_detail', pk=procurement.pk)
    if request.method == 'POST':
        messages.success(request, f"ProcurementPlan {procurement} has been trashed successfully.")
        procurement.delete()
        return redirect('requisition_list')
    return render(request, 'delete_confirmation.html',
                  {'delete': procurement, "cancel_url": reverse('requisition_list')})


def is_programs_head(user):
    return Department.objects.filter(
        name__iexact='Programs',
        head=user
    ).exists()


def is_executive_director(user):
    return user.groups.filter(name='ED').exists()


@login_required(login_url="login")
def review_procurement(request, pk):

    procurement = get_object_or_404(ProcurementPlan, pk=pk)

    if not is_programs_head(request.user):
        messages.warning(request,"Only Programs Head can review procurement plans.")
        return redirect("dashboard")

    action = request.POST.get("action")

    if action not in ["Reviewed", "Rejected"]:
        messages.error(request, "Invalid action.")
        return redirect("requisition_detail", pk=procurement.pk)

    if action == "Reviewed":

        try:
            procurement.validate_budget_limit()
        except ValidationError as error:
            add_validation_messages(request, error)
            return redirect("requisition_detail", pk=procurement.pk)

        procurement.reviewed_by = request.user
        procurement.reviewed_at = timezone.now()
        procurement.status = "Reviewed"

        procurement.save(update_fields=["reviewed_by", "reviewed_at", "status",])

        _notify_next_stage(request, procurement,"ED")
        messages.success(request,"Procurement Plan reviewed successfully.")

    else:
        reason = request.POST.get("reason")
        procurement.rejection_reason = reason
        procurement.rejected_by = request.user
        procurement.reviewed_by = procurement.reviewed_at = None
        procurement.rejected_at = 'reh'
        procurement.rejected_at = timezone.now()
        procurement.status = "Rejected"

        procurement.save(
            update_fields=[
                "rejected_by",
                "rejected_at",
                "rejection_reason",
                "status",
            ]
        )

        _notify_requester(request,
            procurement,
            f"""
                        Procurement Plan {procurement.number} was rejected by Programs Head.
                        Reason: {reason}
                    """
        )

        messages.success(
            request,
            "Procurement Plan rejected."
        )

    return redirect(
        "requisition_detail",
        pk=procurement.pk
    )


@login_required(login_url="login")
def approve_procurement(request, pk):

    procurement = get_object_or_404(
        ProcurementPlan,
        pk=pk
    )

    if not is_executive_director(request.user):
        messages.warning(
            request,
            "Only Executive Director can approve."
        )
        return redirect("dashboard")

    if not procurement.reviewed_by:
        messages.warning(request, "Programs Head review is required first.")
        return redirect("requisition_detail", pk=pk)

    action = request.POST.get("action")

    if action not in ["Approved", "Rejected"]:
        messages.error(request, "Invalid action.")
        return redirect("requisition_detail", pk=pk)

    if action == "Approved":

        try:
            procurement.validate_budget_limit()
        except ValidationError as error:
            add_validation_messages(request, error)
            return redirect("requisition_detail", pk=pk)

        procurement.approved_by = request.user
        procurement.approved_at = timezone.now()
        procurement.status = "Approved"

        procurement.save(
            update_fields=["approved_by", "approved_at", "status",])

        _notify_requester(request,
            procurement, f"Your Procurement Plan {procurement.number} has been approved.")

        messages.success(request, "Procurement Plan approved successfully.")

    else:

        reason = request.POST.get("reason")
        procurement.rejection_reason = reason
        procurement.rejected_by = request.user
        procurement.rejected_at = timezone.now()
        procurement.status = "Rejected"
        procurement.reviewed_by = procurement.reviewed_at = None

        procurement.save(
            update_fields=[
                "rejected_by",
                "rejected_at",
                "rejection_reason",
                "status",
            ]
        )

        _notify_requester(request,
            procurement,
            f"""
                        Procurement Plan {procurement.number} was rejected by the Executive Director.
                        Reason: {reason}
                    """)
        notify_programs_head(request, procurement)

        messages.success(
            request,
            "Procurement Plan rejected."
        )

    return redirect(
        "requisition_detail",
        pk=pk
    )


def _notify_next_stage(request, req, group_name):

    try:
        group = Group.objects.get(name=group_name)
    except Group.DoesNotExist:
        return

    users = group.user_set.all()

    for user in users:

        # In-app notification
        action_url = request.build_absolute_uri(
            reverse("requisition_detail", args=[req.pk]))

        notify(
            request=request,
            users=user,
            title="Procurement Plan Approval",
            message=(
                f"Dear {user},\n\n"
                f"the procurement plan with {req.number} has been \n"
                f"{req.status}.\n\n"
                f"Please check and confirm the approval of the procurement plan \n."
            ),
            category="info",
            source_app="procurement",
            action_url=action_url if req else None,
        )


        # Email notification
        try:
            if user.email:
                send_html_email(
                    request=request,
                    subject="Procurement Plan Approval Required",
                    recipient=user.email,
                    title=f"Procurement Plan ({req.number}) requires your approval",
                    message="Follow the link below to review and approve this request.",
                    relative_link=action_url,
                )
        except Exception as e:
            print(f"Email failed for {user.username}: {e}")

    messages.success(
        request,
        f"Notifications sent to {users.count()} member(s) of {group_name}."
    )


def _notify_requester(request, req, message):
    action_url = request.build_absolute_uri(
        reverse("requisition_detail", args=[req.pk]))

    notify(
        request=request,
        users=req.requester,
        title="Procurement Plan Approval",
        message=(
            f"Dear {req.requester},\n\n"
            f"the procurement plan with {req.number} has been \n"
            f"{req.status}.\n\n"
            f"Please confirm."
        ),
        category="info",
        source_app="procurement",
        action_url=action_url,
    )
    try:
        send_mail(
            subject="ProcurementPlan Update",
            message=f'{message} \n locate at: {action_url}',
            from_email=settings.EMAIL_HOST_USER,
            recipient_list=[req.requester.email],
            fail_silently=True
        )
    except:
        pass



@login_required(login_url='login')
def spent_report(request):
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=90)
    date_from = parse_date_safe(request.GET.get('start'))
    date_to = parse_date_safe(request.GET.get('end'))
    export_type = request.GET.get('export')

    if date_from:
        start_date = date_from
    if date_to:
        end_date = date_to

    data_qs = (Supplier.objects
               .filter(purchase_orders__sent=True, purchase_orders__issue_date__range=[start_date, end_date])
               .values('id', 'full_name', 'title')
               .annotate(total_spent=Sum('purchase_orders__final_amount'))
               .order_by('-total_spent'))

    data = list(data_qs)
    grand_total = sum([d['total_spent'] or 0 for d in data])

    context = {
        'data': data,
        'grand_total': grand_total,
        'start_date': start_date,
        'end_date': end_date,
        'now': timezone.now(),
    }

    if export_type == 'pdf':
        return render_pdf('procurement/spent_report_pdf.html', context, 'spent_report.pdf')

    if export_type == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Spent Report"
        ws.append(['Supplier', 'Total Spent (UXG)'])
        for d in data:
            supplier_name = d.get('full_name') or d.get('title') or 'Unknown'
            ws.append([supplier_name, d['total_spent']])
        ws.append(['Grand Total', grand_total])
        for i, col in enumerate(ws.columns, 1):
            ws.column_dimensions[get_column_letter(i)].width = 30
        resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename=spent_report.xlsx'
        wb.save(resp)
        return resp

    return render(request, 'procurement/spent_report.html', context)


@login_required(login_url='login')
def supplier_report(request, supplier_id=None):
    suppliers = Supplier.objects.all().order_by('full_name')
    supplier = None
    po_list = []
    supplier_total = 0
    export_type = request.GET.get('export')

    supplier_data_qs = (
        PurchaseOrder.objects
        .filter(sent=True)
        .values(
            'supplier__id',
            'supplier__full_name',
            'supplier__title'
        )
        .annotate(total=Sum('final_amount'))
        .order_by('-total')
    )

    supplier_data = list(supplier_data_qs)

    if supplier_id:
        supplier = get_object_or_404(Supplier, pk=supplier_id)
        # show POs for plans that the supplier was sent RFQs for
        # (get plans that have an RFQ where supplier was included)
        po_list = Requisition.objects.filter(purchase_orders__supplier=supplier, purchase_orders__sent=True).distinct().order_by('-date')
        supplier_total = po_list.aggregate(total=Sum('purchase_orders__final_amount'))['total'] or 0

    grand_total = Supplier.objects.aggregate(total=Sum('purchase_orders__final_amount'))['total'] or 0

    context = {
        'supplier_data': supplier_data,
        'suppliers': suppliers,
        'supplier': supplier,
        'po_list': po_list,
        'supplier_total': supplier_total,
        'grand_total': grand_total,
        'now': timezone.now(),
    }

    if export_type == 'pdf':
        return render_pdf('procurement/supplier_report_pdf.html', context, 'supplier_report.pdf')

    if export_type == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Supplier Report"

        logo_url = 'static/assets/images/ricewn.png'
        try:
            logo = Image(logo_url)
            logo.height = 80
            logo.width = 210
            ws.add_image(logo, 'A1')
        except FileNotFoundError:
            print('Logo not found.')

        if supplier:
            title_row = 6
            ws.merge_cells(f'A{title_row}:c{title_row}')
            ws[f'A{title_row}'] = f'{supplier} Report'.upper()
            ws[f'A{title_row}'].font = Font(bold=True, size=14)
            ws[f'A{title_row}'].alignment = Alignment(horizontal='center', vertical='center')

            ws.append(['PO Number', 'Date', 'Total (UXG)'])
            for po in po_list:
                ws.append([po.number, po.date.strftime("%Y-%m-%d"), po.total])
            ws.append(['', f'Total for {supplier.full_name or supplier.title}', supplier_total])
        else:
            ws.append(['Supplier', 'Total Spent (₦)'])
            for item in supplier_data:
                supplier_name = item.get('full_name') or item.get('title')
                ws.append([supplier_name, item['total']])
            ws.append(['Grand Total', grand_total])
        for i, col in enumerate(ws.columns, 1):
            ws.column_dimensions[get_column_letter(i)].width = 30
        resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename=supplier_report.xlsx'
        wb.save(resp)
        return resp

    return render(request, 'procurement/supplier_report.html', context)


@login_required(login_url='login')
def supplier_spend_report_detail(request, pk):

    purchase_order = get_object_or_404(
        PurchaseOrder.objects.select_related('supplier', 'requisition'),
        pk=pk, sent=True,)
    report, _ = SupplierSpendReport.objects.get_or_create(
        purchase_order=purchase_order,
        defaults={
            'supplier': purchase_order.supplier,
            'requisition': purchase_order.requisition,
        }
    )
    if report.supplier_id != purchase_order.supplier_id or report.requisition_id != purchase_order.requisition_id:
        report.supplier = purchase_order.supplier
        report.requisition = purchase_order.requisition
        report.save(update_fields=['supplier', 'requisition', 'last_updated'])

    items = []

    if report.requisition:
        items = report.requisition.items.all() if report.requisition else []
    context = {'report': report, 'items': items,}

    return render(
        request,
        'procurement/supplier_spend_report_detail.html',
        context
    )


@login_required(login_url='login')
def supplier_spend_report_update(request, pk):

    report = get_object_or_404(SupplierSpendReport, pk=pk)

    cash_req = CashRequisition.objects.filter(
        purchase_order=report.purchase_order,
        status='approved',
    ).first()
    if not cash_req:
        messages.info(request, 'Ensure a cash requisition is linked to this LPO and approved before proceeding.')
        return redirect('supplier_spend_report_detail', pk=report.purchase_order_id)

    if report.invoice_no and report.payment_status == "Paid":
        messages.warning(request, 'This invoice has been updated previously and payment made.')
        return redirect('supplier_spend_report_detail', pk=report.purchase_order.id)

    if request.method == 'POST':
        form = SupplierSpendReportUpdateForm(request.POST, instance=report)

        if form.is_valid():
            report = form.save(commit=False)
            report.invoice_amount = user_amount_to_ugx(form.cleaned_data['invoice_amount'], request)
            report.save()

            messages.success(request,'Supplier invoice details updated successfully.')

            return redirect('supplier_spend_report_detail', pk=report.purchase_order_id)

    else:
        form = SupplierSpendReportUpdateForm(
            instance=report,
            initial={
                'invoice_amount': display_amount_from_ugx(report.invoice_amount, request),
            },
        )

    form.fields['invoice_amount'].label = f"Invoice Amount ({get_user_currency(request)})"

    context = {'report': report, 'form': form, 'cash_req': cash_req}

    return render(request,'procurement/supplier_spend_report_update.html',context)


@head_of_procurement_required
def trash_product(request, pk, name):
    pro = get_object_or_404(Product, pk=pk, name=name)
    if request.method == 'POST':
        messages.success(request, f"{pro} has been trashed successfully.")
        pro.delete()
        return redirect('product_list')
    return render(request, 'delete_confirmation.html',
                  {'delete': pro, "cancel_url": reverse('product_list')})


class ProductListView(ListView):
    model = Product
    template_name = "product/product_list.html"
    paginate_by = 10
    context_object_name = 'products'
    ordering = ['-category']


@head_of_procurement_required
def update_product(request, name, pk):
    pro = get_object_or_404(Product, name=name, pk=pk)

    if request.method == 'POST':
        form = ProductForm(request.POST or None, instance=pro)
        if form.is_valid():
            sup = form.save(commit=False)
            sup.save()
            messages.success(request, f"Product: {pro} updated.")
            return redirect('product_list')
    else:
        form = ProductForm(instance=pro)
    return render(request, 'product/product_form.html',
                  {'form': form, 'is_update':True, 'pro':pro,'title':'Update Product'})


@head_of_procurement_required
def create_product(request):
    if request.method == 'POST':
        form = ProductForm(request.POST or None)
        if form.is_valid():
            form.save()
            messages.success(request, f"Product ({form.cleaned_data['name']}) created successfully.")
            return redirect('product_list')
    else:
        form = ProductForm()
    return render(request, 'product/product_form.html', {'form': form, 'title':'Create Product'})
