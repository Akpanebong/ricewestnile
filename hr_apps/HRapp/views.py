from email.mime.image import MIMEImage

from django.conf import settings
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.core.mail import send_mail
from django.core.paginator import Paginator
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views.generic import DetailView, UpdateView
from account.models import Profile, Unit
from notification.utils import notify
from .models import Employee, Attendance, Leave, \
    Training, ForumThread, ForumPost, SituationReport, LeaveType
from django.contrib.auth.decorators import login_required,  user_passes_test
from .forms import EmployeeForm, LeaveForm, AttendanceForm, TrainingForm, \
    ForumThreadForm, ForumPostForm, SituationReportForm, UpdateSitrepStatusForm,\
    StaffAttendanceForm
from django.contrib import messages
from django.http import HttpResponse, HttpResponseForbidden
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from django.utils.dateparse import parse_date
import os
from openpyxl.styles import Font, Alignment
from openpyxl import Workbook
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.core.mail import EmailMultiAlternatives
from datetime import date, datetime
from django.utils.timezone import now, timedelta
from django.db.models import Avg, Count, Q
from .utils import (render_to_pdf, yearly_leave_allocation_report, staff_leave_balance_report, employee_leave_balances,
                    logo_path, )
from account.permissions import is_cmt as is_cmt_user
import logging

logger = logging.getLogger(__name__)


# helper access checks (customize to your groups/permissions)
def is_supervisor(user):
    return bool(user and (user.groups.filter(name="Supervisor").exists() or user.is_staff))


# def is_hr(user):
#     return user.groups.filter(name="HR").exists()
def is_hr(user):
    return (
        bool(user)
        and (
            user.is_superuser
            or (
                getattr(user, "profile_type", None) == "Staff"
                and user.groups.filter(name="HR").exists()
            )
        )
    )


def is_ed(user):
    return bool(user and (user.groups.filter(name="ED").exists() or user.is_superuser))



def is_cmt(user):
    return is_cmt_user(user)



def is_unit_head(user):
    return bool(user and Unit.objects.filter(head=user).exists())


@login_required(login_url="login")
def dashboard(request):
    from hr_apps.appraisals.models import Appraisal, PerformanceImprovementPlan
    from hr_apps.vacancy.models import Applicant, JobOpening, RecruitmentRequest

    user = request.user
    is_admin_view = is_supervisor(user) or is_hr(user) or is_ed(user) or is_cmt(user)
    if not is_admin_view:
        return redirect('update_employee', user.slug)

    today = now().date()
    current_year = today.year
    last_7_days = [today - timedelta(days=i) for i in range(6, -1, -1)]

    # ==========================
    # CORE KPIs
    # ==========================
    total_employees = Employee.objects.exclude(user__status__in=["Exit", "Suspended"]).count()

    pending_leave_count = Leave.objects.filter(status="Pending").count()

    active_today = Attendance.objects.filter(
        date__date=today,
        status="Present"
    ).count()

    trainings_this_month = Training.objects.filter(
        start_date__month=today.month
    ).count()

    open_jobs = JobOpening.objects.filter(is_active=True).filter(
        Q(deadline__isnull=True) | Q(deadline__gte=today)
    ).count()

    pending_recruitment_requests = RecruitmentRequest.objects.filter(status="Pending").count()
    pending_applicants = Applicant.objects.filter(status="Pending").count()

    total_appraisals = Appraisal.objects.count()
    appraisals_pending_hr = Appraisal.objects.filter(status="reviewed_by_supervisor").count()
    appraisals_pending_cmt = Appraisal.objects.filter(status="reviewed_by_hr").count()
    pips_pending_staff = PerformanceImprovementPlan.objects.filter(status="sent_to_staff").count()
    average_job_score = Appraisal.objects.aggregate(avg=Avg("job_description_score"))["avg"] or 0

    pending_leaves = (Leave.objects.select_related("employee").filter(status="Pending")
                      .order_by("-applied_date")[:6] if is_admin_view else
                      Leave.objects.filter(employee=user))

    dept_data = (Profile.objects.select_related("unit").values("unit__name")
                 .annotate(total=Count("id")).order_by("unit__name"))

    department_labels = [d["unit__name"] for d in dept_data]
    department_values = [d["total"] for d in dept_data]
    print(department_labels, department_values)

    leave_status_data = (
        Leave.objects
        .values("status")
        .annotate(total=Count("id"))
    )

    leave_labels = [l["status"] for l in leave_status_data]
    leave_values = [l["total"] for l in leave_status_data]

    # ==========================
    # ATTENDANCE TREND (7 DAYS)
    # ==========================
    attendance_counts = {
        d: Attendance.objects.filter(
            date__date=d,
            status="Present"
        ).count()
        for d in last_7_days
    }

    attendance_labels = [d.strftime("%a") for d in last_7_days]
    attendance_values = list(attendance_counts.values())

    # ==========================
    # TRAINING PARTICIPATION
    # ==========================
    training_data = (
        Training.objects
        .annotate(participants_count=Count("participants"))
        .order_by("-participants_count")
    )

    training_labels = [t.title for t in training_data]
    training_values = [t.participants_count for t in training_data]

    appraisal_status_data = Appraisal.objects.values("status").annotate(total=Count("id")).order_by("status")
    appraisal_labels = [row["status"] for row in appraisal_status_data]
    appraisal_values = [row["total"] for row in appraisal_status_data]

    recruitment_status_data = RecruitmentRequest.objects.values("status").annotate(total=Count("id")).order_by("status")
    recruitment_labels = [row["status"] for row in recruitment_status_data]
    recruitment_values = [row["total"] for row in recruitment_status_data]

    recent_recruitment_requests = (
        RecruitmentRequest.objects
        .select_related("unit", "requested_by")
        .order_by("-created_at")[:6]
    )

    recent_appraisals = (
        Appraisal.objects
        .select_related("employee__user", "supervisor__user")
        .order_by("-created_at")[:6]
    )

    return render(request, "hr/dashboard.html", {
        # KPI
        "total_employees": total_employees,
        "pending_leaves": pending_leave_count,
        "active_today": active_today,
        "trainings_this_month": trainings_this_month,
        "open_jobs": open_jobs,
        "pending_recruitment_requests": pending_recruitment_requests,
        "pending_applicants": pending_applicants,
        "total_appraisals": total_appraisals,
        "appraisals_pending_hr": appraisals_pending_hr,
        "appraisals_pending_cmt": appraisals_pending_cmt,
        "pips_pending_staff": pips_pending_staff,
        "average_job_score": average_job_score,
        "current_year": current_year,

        # Lists
        "pending_leave_list": pending_leaves,
        "recent_recruitment_requests": recent_recruitment_requests,
        "recent_appraisals": recent_appraisals,

        # Charts
        "department_labels": department_labels,
        "department_data": department_values,
        "leave_labels": leave_labels,
        "leave_data": leave_values,
        "attendance_labels": attendance_labels,
        "attendance_data": attendance_values,
        "training_labels": training_labels,
        "training_data": training_values,
        "appraisal_labels": appraisal_labels,
        "appraisal_data": appraisal_values,
        "recruitment_labels": recruitment_labels,
        "recruitment_data": recruitment_values,
    })


@login_required(login_url='login')
def employee_list(request):
    if not (is_hr(request.user) or is_supervisor(request.user) or request.user.is_superuser):
        return HttpResponseForbidden("Access denied.")

    staff_qs = Profile.objects.filter(
        profile_type__in=["Staff", "Intern", 'Volunteer', 'Community ...']
    ).select_related("department").order_by("username")

    paginator = Paginator(staff_qs, 10)
    page_number = request.GET.get("page")
    employees = paginator.get_page(page_number)

    return render(request, "hr/employee_list.html", {
        "employees": employees
    })


@login_required(login_url='login')
def admin_edit_employee(request, pk, slug):
    employee = get_object_or_404(Employee, pk=pk, slug=slug)
    if not (is_supervisor(user=request.user) or is_hr(request.user)):
        messages.error(request, 'Access Denied.')
        return redirect('logout')

    if request.method == 'POST':
        form = EmployeeForm(request.POST, instance=employee)
        if form.is_valid():
            form.save()
            messages.success(request, "Employee updated successfully.")
            return redirect('employee_list')
    else:
        form = EmployeeForm(instance=employee)
    return render(request, 'hr/edit_employee.html', {'form': form, 'employee': employee})



@login_required(login_url='login')
def delete_employee(request, pk, slug):
    employee = get_object_or_404(Employee, pk=pk, slug=slug)

    if not (is_supervisor(user=request.user) or is_hr(request.user)):
        messages.error(request, 'Access Denied.')
        return redirect('logout')

    if request.method == 'POST':
        employee.delete()
        messages.success(request, "Employee deleted successfully.")
        return redirect('employee_list')
    return render(request, 'hr/delete_confirmation.html', {'delete': employee,
                                                           "cancel_url": reverse('employee_list')})


# Add Employee
@login_required(login_url='login')
def add_employee(request):
    if not (is_supervisor(user=request.user) or is_hr(request.user)):
        messages.error(request, 'Access Denied.')
        return redirect('logout')

    if request.method == 'POST':
        form = EmployeeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('employee_list')
    else:
        form = EmployeeForm()
    return render(request, 'hr/add_employee.html', {'form': form})


@login_required(login_url="login")
def leave_list(request):
    user = request.user
    employee = getattr(user, "employee", None)

    # Executive Director
    if is_ed(user):
        leaves_qs = Leave.objects.all().order_by("-applied_date")

    # HR
    elif is_hr(user):
        leaves_qs = Leave.objects.filter(
            supervisor_approved=True
        ).order_by("-applied_date")

    # Supervisor
    elif is_supervisor(user):
        if employee is None:
            messages.error(request, "Supervisor employee record not found.")
            return redirect("dashboard")

        leaves_qs = Leave.objects.filter(
            supervisor=employee
        ).order_by("-applied_date")

    # Employee
    else:
        if employee is None:
            messages.error(request, "Employee profile not found.")
            return redirect("dashboard")

        leaves_qs = Leave.objects.filter(
            employee=user
        ).order_by("-end_date")

    paginator = Paginator(leaves_qs, 5)
    page_number = request.GET.get("page")

    return render(request, "hr/leave_list.html", {
        "leaves": paginator.get_page(page_number),
    })


@login_required(login_url='login')
def leave_detail(request, pk, slug):
    leave = get_object_or_404(Leave, pk=pk, slug=slug)
    employee = getattr(request.user, "employee", None)
    can_view = (
        request.user.is_superuser
        or leave.employee_id == request.user.id
        or (employee and leave.supervisor_id == employee.id)
        or is_hr(request.user)
        or is_ed(request.user)
    )
    if not can_view:
        return HttpResponseForbidden("Access denied.")
    return render(request, "hr/leave_detail.html", {"leave": leave})


@login_required(login_url='login')
def apply_leave(request):
    if request.user.status == "On Leave" or request.user.status == "Probation":
        messages.info(request, f"Oops!!!, Your profile is still on {request.user.status}, so you can't apply now")
        return redirect('leave_list')

    year = date.today().year
    leave_types = list(LeaveType.objects.all().order_by("name"))
    default_leave_type = leave_types[0] if leave_types else None
    selected_leave_type_id = (
        request.POST.get("leave_type")
        or request.GET.get("leave_type")
        or (str(default_leave_type.id) if default_leave_type else "")
    )

    balance_rows = employee_leave_balances(employee=request.user, year=year)
    leave_type_meta = {
        str(row["leave_type"].id): {
            "id": int(row["leave_type"].id),
            "name": str(row["leave_type"].name),
            "allotted_days": int(row["allotted_days"] or 0),
            "taken_days": int(row["taken_days"] or 0),
            "remaining_days": int(row["remaining_days"] or 0),
            "is_special": bool(getattr(row["leave_type"], "is_special", False)),
        }
        for row in balance_rows
    }

    selected_meta = leave_type_meta.get(str(selected_leave_type_id))
    if selected_meta is None and default_leave_type is not None:
        selected_meta = leave_type_meta.get(str(default_leave_type.id))
        selected_leave_type_id = str(default_leave_type.id)

    try:
        if request.user.profile_type: # == "Staff":
            employee = request.user.employee

    except Employee.DoesNotExist:
        messages.error(request, "You don't have an employee profile. Contact HR.")
        return redirect("dashboard")

    supervisor = request.user.employee.supervised_by # if request.user.profile_type == "Staff" else intern.supervised_by
    if not supervisor:
        messages.error(request, "No supervisor assigned. Contact HR.")
        return redirect("dashboard")

    if request.method == "POST":
        form = LeaveForm(request.POST)
        # Ensure model-level validation has the employee attached during form.is_valid()
        form.instance.employee = request.user
        form.instance.supervisor = supervisor

        if form.is_valid():
            leave = form.save(commit=False)
            leave.employee = request.user
            leave.supervisor = supervisor

            leave.save()
            notify(request=request, users=supervisor.user, title=f"{leave.leave_type} Leave Application",
                   message=(f"Dear {supervisor},\n\n "
                            f"A leave application has been submitted by {leave.employee}. \n\n "
                            f"Please review and take action on this application. "
                            ),category="info", source_app="hr",
                   action_url=request.build_absolute_uri(
                       reverse("leave_detail", kwargs={"pk": leave.pk, "slug": leave.slug})
                   ),)

            _send_html_email(
                subject=f"{leave.leave_type} Leave Application",
                to=[supervisor.user.email],
                template="hr/email/leave_application_notification.html",
                context={
                    "leave": leave,
                    "review_url": request.build_absolute_uri(
                        reverse("leave_detail",kwargs={"pk": leave.pk, "slug": leave.slug,},
                        )
                    ),
                },
            )
            messages.success(request, "Leave applied successfully. "
                                      "\nYou leave will be shown here once your supervisor has approve")
            return redirect("leave_list")
        else:
            errors = []
            for field, field_errors in form.errors.items():
                for err in field_errors:
                    if field == "__all__":
                        errors.append(str(err))
                    else:
                        label = getattr(form.fields.get(field), "label", field)
                        errors.append(f"{label}: {err}")
            if errors:
                messages.error(request, " ".join(errors))

    else:
        form = LeaveForm(initial={"leave_type": selected_leave_type_id} if selected_leave_type_id else None)
        form.instance.employee = request.user
        form.instance.supervisor = supervisor

    return render(request, "hr/apply_leave.html", {
        "form": form,
        "year": year,
        "leave_types": leave_types,
        "selected_leave_type_id": str(selected_leave_type_id or ""),
        "selected_leave_type_name": (selected_meta or {}).get("name", ""),
        "selected_allotted_days": int((selected_meta or {}).get("allotted_days", 0) or 0),
        "selected_taken_days": int((selected_meta or {}).get("taken_days", 0) or 0),
        "selected_remaining_days": int((selected_meta or {}).get("remaining_days", 0) or 0),
        "leave_type_meta": leave_type_meta,
    })


@login_required(login_url='login')
@user_passes_test(is_supervisor)
def approve_supervisor(request, pk, action):
    leave = get_object_or_404(Leave, pk=pk)

    if request.method == "POST":

        if action == "approved":
            leave.supervisor = request.user.employee
            leave.supervisor_approved = True
            leave.supervisor_approved_at = timezone.now()

        elif action == "denied":
            leave.status = "Denied"
            leave.employee.status = "Active"
            leave.employee.save(update_fields=["status"])
            leave.save(update_fields=["status"])
            messages.error(request, "Leave rejected.")
            return redirect(reverse("leave_detail", kwargs={"pk": pk, "slug": leave.slug}))

        elif action == "reverse":
            leave.supervisor_approved = False

        else:
            messages.warning(request, "Invalid action.")
            return redirect(reverse("leave_detail", kwargs={"pk": pk, "slug": leave.slug}))

        leave.update_status()
        leave.save()
        messages.success(request, "Supervisor action recorded.")
        return redirect(reverse("leave_detail", kwargs={"pk": pk, "slug": leave.slug}))

    return render(request, "hr/confirm_action.html", {"obj": leave, "action": action})


@login_required(login_url='login')
@user_passes_test(is_hr)
def approve_hr(request, pk, action):
    leave = get_object_or_404(Leave, pk=pk)

    if request.method == "POST":
        leave.hr = request.user

        if action == "approved":
            leave.hr_approved = True
            leave.hr_approved_at = timezone.now()

        elif action == "denied":
            leave.status = "Denied"
            leave.employee.status = "Active"
            leave.employee.save(update_fields=["status"])
            leave.save(update_fields=["status"])
            messages.error(request, "Leave rejected.")
            return redirect(reverse("leave_detail", kwargs={"pk": pk, "slug": leave.slug}))

        elif action == "reverse":
            if leave.ed_approved:
                messages.error(request, "Cannot reverse after ED approval.")
                return redirect(reverse("leave_detail", kwargs={"pk": pk, "slug": leave.slug}))

            leave.hr_approved = False

        leave.update_status()
        leave.save()
        messages.success(request, "HR action recorded.")
        return redirect(reverse("leave_detail", kwargs={"pk": pk, "slug": leave.slug}))

    return render(request, "hr/confirm_action.html", {"obj": leave, "action": action})


@login_required(login_url='login')
@user_passes_test(is_ed)
def approve_ed(request, pk, action):
    leave = get_object_or_404(Leave, pk=pk)

    if request.method == "POST":
        leave.ed = request.user

        if action == "approved":
            leave.ed_approved = True
            leave.employee.status = "On Leave"
            leave.ed_approved_at = timezone.now()
            leave.employee.save(update_fields=["status"])

        elif action == "denied":
            leave.status = "Denied"
            leave.employee.status = "Active"
            leave.employee.save(update_fields=["status"])
            leave.save(update_fields=["status"])
            messages.error(request, "Leave rejected.")
            return redirect(reverse("leave_detail", kwargs={"pk": pk, "slug": leave.slug}))

        elif action == "reverse":
            if leave.ed_approved:
                messages.error(request, "Already finalized. Cannot reverse.")
                return redirect(reverse("leave_detail", kwargs={"pk": pk, "slug": leave.slug}))

            leave.ed_approved = False

        leave.update_status()
        leave.save()
        messages.success(request, "ED action recorded.")
        return redirect(reverse("leave_detail", kwargs={"pk": pk, "slug": leave.slug}))

    return render(request, "hr/confirm_action.html", {"obj": leave, "action": action})


@login_required(login_url='login')
def delete_leave(request, pk):
    if not (is_supervisor(user=request.user) or is_hr(request.user)):
        messages.error(request, 'Access Denied.')
        return redirect('logout')

    leave = get_object_or_404(Leave, pk=pk)
    if request.method == 'POST':
        leave.employee.status = 'Active'

        leave.employee.save()
        leave.delete()
        messages.success(request, "Leave deleted successfully.")
        return redirect('leave_list')
    return render(request, 'hr/delete_confirmation.html', {'delete': leave,
                                                           "cancel_url": reverse('leave_list')})



@login_required(login_url="login")
def leave_yearly_report(request):
    if not (is_supervisor(request.user) or is_hr(request.user) or is_ed(request.user) or request.user.is_superuser):
        return HttpResponseForbidden("Access denied.")

    year = int(request.GET.get("year", date.today().year))
    start_date = parse_date(request.GET.get("start_date") or "")
    end_date = parse_date(request.GET.get("end_date") or "")

    leave_types = list(LeaveType.objects.all().order_by("name"))
    default_leave_type = leave_types[0] if leave_types else None
    leave_type_id = request.GET.get("leave_type") or (str(default_leave_type.id) if default_leave_type else "")
    selected_leave_type = next((lt for lt in leave_types if str(lt.id) == str(leave_type_id)), default_leave_type)

    report_rows = yearly_leave_allocation_report(year=year, start_date=start_date, end_date=end_date)
    staff_rows = (
        staff_leave_balance_report(
            year=year,
            leave_type=selected_leave_type,
            start_date=start_date,
            end_date=end_date,
            include_zero=True,
        )
        if selected_leave_type
        else []
    )

    # Organization totals (leave type summary)
    total_staff = report_rows[0]["staff_count"] if report_rows else 0
    total_allotted = sum(r["allotted_total_days"] for r in report_rows)
    total_approved = sum(r["approved_total_days"] for r in report_rows)
    total_remaining = sum(r["remaining_total_days"] for r in report_rows)

    # Staff totals (selected leave type)
    staff_count = len(staff_rows)
    staff_allotted_per = int(getattr(selected_leave_type, "no_of_days", 0) or 0) if selected_leave_type else 0
    staff_total_allotted = staff_count * staff_allotted_per
    staff_total_taken = sum(int(r.get("taken_days") or 0) for r in staff_rows)
    staff_total_remaining = max(0, staff_total_allotted - staff_total_taken)

    context = {
        "year": year,
        "start_date": request.GET.get("start_date") or "",
        "end_date": request.GET.get("end_date") or "",

        # org summary
        "report_rows": report_rows,
        "total_staff": total_staff,
        "total_allotted": total_allotted,
        "total_approved": total_approved,
        "total_remaining": total_remaining,

        # staff summary
        "leave_types": leave_types,
        "selected_leave_type": selected_leave_type,
        "staff_rows": staff_rows,
        "staff_total_allotted": staff_total_allotted,
        "staff_total_taken": staff_total_taken,
        "staff_total_remaining": staff_total_remaining,
    }

    return render(request, "hr/leave_report.html", context)


@login_required(login_url="login")
def leave_yearly_report_pdf(request):
    if not (is_supervisor(request.user) or is_hr(request.user) or is_ed(request.user) or request.user.is_superuser):
        return HttpResponseForbidden("Access denied.")

    year = int(request.GET.get("year", date.today().year))
    start_date = parse_date(request.GET.get("start_date") or "")
    end_date = parse_date(request.GET.get("end_date") or "")

    report_rows = yearly_leave_allocation_report(year=year, start_date=start_date, end_date=end_date)

    total_staff = report_rows[0]["staff_count"] if report_rows else 0
    total_allotted = sum(r["allotted_total_days"] for r in report_rows)
    total_approved = sum(r["approved_total_days"] for r in report_rows)
    total_remaining = sum(r["remaining_total_days"] for r in report_rows)

    context = {
        "year": year,
        "start_date": request.GET.get("start_date") or "",
        "end_date": request.GET.get("end_date") or "",
        "report_rows": report_rows,
        "total_staff": total_staff,
        "total_allotted": total_allotted,
        "total_approved": total_approved,
        "total_remaining": total_remaining,
    }

    return render_to_pdf("hr/leave_report_pdf.html", context)


@login_required(login_url="login")
def export_leave_report_excel(request):
    if not (is_supervisor(request.user) or is_hr(request.user) or is_ed(request.user) or request.user.is_superuser):
        return HttpResponseForbidden("Access denied.")

    year = int(request.GET.get("year", date.today().year))
    start_date = parse_date(request.GET.get("start_date") or "")
    end_date = parse_date(request.GET.get("end_date") or "")

    report_rows = yearly_leave_allocation_report(year=year, start_date=start_date, end_date=end_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "Leave Summary"

    ws.merge_cells("A1:E1")
    ws["A1"] = "Leave Allocation Summary"
    ws["A1"].font = Font(bold=True, size=14, color="228B22")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:E2")
    range_label = f"{year}"
    if request.GET.get("start_date") and request.GET.get("end_date"):
        range_label = f"{year} ({request.GET.get('start_date')} to {request.GET.get('end_date')})"
    ws["A2"] = range_label
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(italic=True, size=10)

    headers = ["Leave Type", "Staff Count", "Allotted (Total)", "Taken/Approved", "Remaining (Total)"]
    ws.append(headers)
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.font = Font(bold=True, color="228B22")
        cell.alignment = Alignment(horizontal="center")

    for row in report_rows:
        ws.append(
            [
                row["leave_type"].name,
                int(row["staff_count"] or 0),
                int(row["allotted_total_days"] or 0),
                int(row["approved_total_days"] or 0),
                int(row["remaining_total_days"] or 0),
            ]
        )

    total_staff = report_rows[0]["staff_count"] if report_rows else 0
    total_allotted = sum(r["allotted_total_days"] for r in report_rows)
    total_approved = sum(r["approved_total_days"] for r in report_rows)
    total_remaining = sum(r["remaining_total_days"] for r in report_rows)

    ws.append([])
    ws.append(["TOTAL", total_staff, total_allotted, total_approved, total_remaining])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="leave_report_{year}.xlsx"'
    wb.save(response)
    return response


@login_required(login_url="login")
def export_leave_staff_excel(request):
    if not (is_supervisor(request.user) or is_hr(request.user) or is_ed(request.user) or request.user.is_superuser):
        return HttpResponseForbidden("Access denied.")

    year = int(request.GET.get("year", date.today().year))
    start_date = parse_date(request.GET.get("start_date") or "")
    end_date = parse_date(request.GET.get("end_date") or "")

    leave_types = list(LeaveType.objects.all().order_by("name"))
    default_leave_type = leave_types[0] if leave_types else None
    leave_type_id = request.GET.get("leave_type") or (str(default_leave_type.id) if default_leave_type else "")
    selected_leave_type = next((lt for lt in leave_types if str(lt.id) == str(leave_type_id)), default_leave_type)
    if not selected_leave_type:
        return HttpResponse("No leave types configured.", status=400)

    staff_rows = staff_leave_balance_report(
        year=year,
        leave_type=selected_leave_type,
        start_date=start_date,
        end_date=end_date,
        include_zero=True,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Staff Balances"

    ws.merge_cells("A1:D1")
    ws["A1"] = "Staff Leave Balances"
    ws["A1"].font = Font(bold=True, size=14, color="228B22")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:D2")
    range_label = f"{year} - {selected_leave_type.name}"
    if request.GET.get("start_date") and request.GET.get("end_date"):
        range_label = f"{year} - {selected_leave_type.name} ({request.GET.get('start_date')} to {request.GET.get('end_date')})"
    ws["A2"] = range_label
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(italic=True, size=10)

    headers = ["Employee", "Allotted", "Taken", "Remaining"]
    ws.append(headers)
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.font = Font(bold=True, color="228B22")
        cell.alignment = Alignment(horizontal="center")

    for row in staff_rows:
        ws.append(
            [
                str(row["employee"]),
                int(row["allotted_days"] or 0),
                int(row["taken_days"] or 0),
                int(row["remaining_days"] or 0),
            ]
        )

    staff_count = len(staff_rows)
    allotted_per_staff = int(getattr(selected_leave_type, "no_of_days", 0) or 0)
    total_allotted = staff_count * allotted_per_staff
    total_taken = sum(int(r.get("taken_days") or 0) for r in staff_rows)
    total_remaining = max(0, total_allotted - total_taken)

    ws.append([])
    ws.append(["TOTAL", total_allotted, total_taken, total_remaining])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    safe_name = "".join(c if c.isalnum() or c in ("-", "_") else "_" for c in selected_leave_type.name)
    response["Content-Disposition"] = f'attachment; filename="leave_staff_{safe_name}_{year}.xlsx"'
    wb.save(response)
    return response


@login_required(login_url='login')
def attendance_list(request):
    employee = request.user
    # Staff see only their own records; HR/Admin can see all
    if is_supervisor(user=request.user) or is_hr(request.user):
        record_qs = Attendance.objects.all().order_by("-date")
    else:
        record_qs = Attendance.objects.filter(employee=employee).order_by("-date")
    # record_qs = Attendance.objects.all()
    paginator = Paginator(record_qs, 5)  # 10 employees per page
    page_number = request.GET.get('page')
    records = paginator.get_page(page_number)
    return render(request, 'hr/attendance_list.html', {'records': records})


@login_required(login_url='login')
def mark_attendance(request):
    if not (is_supervisor(request.user) or is_hr(request.user)):
        return HttpResponseForbidden("Access denied.")

    if request.method == 'POST':
        form = AttendanceForm(request.POST)
        if form.is_valid():

            employee = form.cleaned_data['employee']
            today = now().date()
            if Attendance.objects.filter(employee=employee, date__date=today).exists():
                messages.warning(request, f"You have already marked attendance today for {employee}.")
                return redirect("attendance_list")

            form.save()
            messages.success(request, f"You have successfully marked attendance today for {employee}.")
            return redirect('attendance_list')
    else:
        form = AttendanceForm()
    return render(request, 'hr/mark_attendance.html', {'form': form})


@login_required(login_url='login')
def staff_mark_attendance(request):
    employee = request.user
    today = now().date()

    # # Prevent double marking
    if Attendance.objects.filter(employee=employee, date__date=today).exists():
        messages.warning(request, "⚠️ Already marked attendance today.")
        return redirect("attendance_list")

    if request.method == "POST":
        # Get device hash directly from POST
        device_hash = request.POST.get("device_hash")
        form = StaffAttendanceForm(request.POST)
        if form.is_valid():
            # Bind device hash first time
            if not employee.device_hash:
                employee.device_hash = device_hash
                employee.save()
                messages.info(request, "📌 Your account has been bound to this device.")

            # Enforce same device
            if employee.device_hash and employee.device_hash != device_hash:
                messages.error(request, "❌ This account is already bound to another device. Contact admin.")
                return redirect("attendance_list")

            # Save attendance (only employee + date + status)
            attendance = form.save(commit=False)
            attendance.employee = employee
            attendance.date = today
            attendance.save()

            messages.success(request, "✅ Attendance marked successfully!")
            return redirect("attendance_list")
        else:
            messages.error(request, "❌ Failed to mark attendance. Please try again.")
    else:
        form = StaffAttendanceForm()

    return render(request, "hr/staff_mark_attendance.html", {"form": form})


@login_required(login_url='login')
def export_attendance_pdf(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    export_all = request.GET.get("export_all")

    user = request.user

    if user.is_superuser:
        attendances = Attendance.objects.all().order_by("date")
        if not export_all and start_date and end_date:
            start_date = parse_date(start_date)
            end_date = parse_date(end_date)
            if start_date and end_date:
                attendances = attendances.filter(date__range=(start_date, end_date))
    else:
        # Regular employee - only their own records
        attendances = Attendance.objects.filter(employee=user).order_by("date")
        if start_date and end_date:
            start_date = parse_date(start_date)
            end_date = parse_date(end_date)
            if start_date and end_date:
                attendances = attendances.filter(date__range=(start_date, end_date))

    # --- PDF generation code remains unchanged below ---
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="attendance_report.pdf"'

    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # --- Header Section ---
    logo_path = os.path.join(settings.STATICFILES_DIRS[0], "assets/images/ricewn.png")
    if os.path.exists(logo_path):
        p.drawImage(logo_path, 40, height - 80, width=160, height=60, mask="auto")

    p.setFont("Helvetica-Bold", 18)
    p.setFillColor(colors.green)
    p.drawString(220, height - 50, "RICE WEST NILE")
    p.setFont("Helvetica", 12)
    p.drawString(240, height - 70, "Attendance Report")

    if not export_all and start_date and end_date:
        p.setFont("Helvetica-Oblique", 10)
        p.setFillColor(colors.black)
        p.drawString(220, height - 85, f"From {start_date} to {end_date}")
    else:
        p.setFont("Helvetica-Oblique", 10)
        p.setFillColor(colors.black)
        p.drawString(260, height - 85, "All Records")

    # --- Table Section ---
    y = height - 140
    row_height = 20

    # Table headers in green
    p.setFont("Helvetica-Bold", 11)
    p.setFillColor(colors.green)
    p.drawString(50, y, "Employee")
    p.drawString(250, y, "Date")
    p.drawString(400, y, "Status")
    y -= row_height
    p.setFont("Helvetica", 10)
    p.setFillColor(colors.black)

    total_present = 0
    total_absent = 0
    total_leave = 0

    for index, att in enumerate(attendances, start=1):
        # Alternate background
        if index % 2 == 0:
            p.setFillColorRGB(0.95, 0.95, 0.95)
            p.rect(45, y - 5, width - 90, row_height, stroke=0, fill=1)
            p.setFillColor(colors.black)

        # Draw row
        p.drawString(50, y, att.employee.get_full_name())
        p.drawString(250, y, att.date.strftime("%Y-%m-%d  %H:%M:%S"))
        p.drawString(400, y, att.status)

        # Count summary
        if att.status.lower() == "present":
            total_present += 1
        elif att.status.lower() == "absent":
            total_absent += 1
        elif att.status.lower() == "leave":
            total_leave += 1

        y -= row_height
        if y < 50:
            p.showPage()
            y = height - 50

    # --- Summary Row ---
    y -= 20
    p.setFont("Helvetica-Bold", 11)
    p.setFillColor(colors.green)
    p.drawString(50, y, "Summary")
    p.setFillColor(colors.black)
    p.setFont("Helvetica", 10)
    p.drawString(150, y, f"Present: {total_present}")
    p.drawString(300, y, f"Absent: {total_absent}")
    p.drawString(450, y, f"Leave: {total_leave}")

    p.save()
    return response


@login_required(login_url='login')
def export_attendance_excel(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    export_all = request.GET.get("export_all")

    user = request.user

    if user.is_superuser:
        attendances = Attendance.objects.all().order_by("date")
        if not export_all and start_date and end_date:
            start_date = parse_date(start_date)
            end_date = parse_date(end_date)
            if start_date and end_date:
                attendances = attendances.filter(date__range=(start_date, end_date))
    else:
        attendances = Attendance.objects.filter(employee=user.employee).order_by("date")
        if start_date and end_date:
            start_date = parse_date(start_date)
            end_date = parse_date(end_date)
            if start_date and end_date:
                attendances = attendances.filter(date__range=(start_date, end_date))

    # --- Excel generation code remains unchanged below ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # --- Header ---
    ws.merge_cells("A1:C1")
    ws["A1"] = "RICE WEST NILE - Attendance Report"
    ws["A1"].font = Font(bold=True, color="228B22", size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    if not export_all and start_date and end_date:
        ws.merge_cells("A2:C2")
        ws["A2"] = f"From {start_date} to {end_date}"
    else:
        ws.merge_cells("A2:C2")
        ws["A2"] = "All Records"

    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(italic=True, size=10)

    # Column headers
    headers = ["Employee", "Date", "Status"]
    ws.append(headers)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.font = Font(bold=True, color="228B22")  # green text
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    total_present = 0
    total_absent = 0
    total_leave = 0

    for att in attendances:
        ws.append([
            att.employee.get_full_name(),
            att.date.strftime("%Y-%m-%d"),
            att.status
        ])

        if att.status.lower() == "present":
            total_present += 1
        elif att.status.lower() == "absent":
            total_absent += 1
        elif att.status.lower() == "leave":
            total_leave += 1

    # Summary row
    last_row = ws.max_row + 2
    ws[f"A{last_row}"] = "Summary"
    ws[f"A{last_row}"].font = Font(bold=True, color="228B22")
    ws[f"B{last_row}"] = f"Present: {total_present}"
    ws[f"C{last_row}"] = f"Absent: {total_absent} | Leave: {total_leave}"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="attendance_report.xlsx"'
    wb.save(response)
    return response



@login_required(login_url='login')
def training_list(request):
    trainings_qs = Training.objects.all().order_by('-start_date')
    paginator = Paginator(trainings_qs, 4)  # 10 employees per page
    page_number = request.GET.get('page')
    trainings = paginator.get_page(page_number)
    return render(request, 'hr/training_list.html', {'trainings': trainings})


class TrainingDetailModalView(LoginRequiredMixin, DetailView):
    model = Training
    template_name = "hr/training_detail_modal.html"
    context_object_name = "training"

    def get_queryset(self):
        return (
            Training.objects
            .prefetch_related("participants", "participants__department")
        )

    def render_to_response(self, context, **response_kwargs):
        # Only allow AJAX access
        if not self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            return HttpResponseForbidden("Invalid request")
        return super().render_to_response(context, **response_kwargs)


@login_required(login_url='login')
def add_training(request):
    if not is_supervisor(user=request.user):
        messages.error(request, 'Access Denied.')
        return redirect('logout')

    if request.method == 'POST':
        form = TrainingForm(request.POST)
        if form.is_valid():
            form.save()
            participants = form.cleaned_data['participants']
            start_date = form.cleaned_data['start_date']
            description = form.cleaned_data['description']
            title = form.cleaned_data['title']
            for employee in participants:
                notify(
                    title=title,
                    message=f'You are invited for a training commencing on {start_date}. {description}',
                    users=employee,
                    action_url=""
                )
                send_mail(subject=title, message=description, from_email=settings.EMAIL_HOST_USER,
                          recipient_list=[employee.email], fail_silently=False)
            return redirect('training_list')
    else:
        form = TrainingForm()
    return render(request, 'hr/add_training.html', {'form': form})


class TrainingUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Training
    form_class = TrainingForm
    template_name = "hr/training_edit.html"
    context_object_name = "training"

    def get_queryset(self):
        return Training.objects.prefetch_related("participants")

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        if obj.slug != self.kwargs.get("slug"):
            from django.http import Http404
            raise Http404("Invalid URL")
        return obj

    def test_func(self):
        return self.request.user.is_staff

    def form_valid(self, form):
        participants = form.cleaned_data['participants']
        start_date = form.cleaned_data['start_date']
        end_date = form.cleaned_data['end_date']
        description = form.cleaned_data['description']
        title = form.cleaned_data['title']
        if end_date < start_date:
            form.add_error("end_date", "End date cannot be before start date.")
            return self.form_invalid(form)

        for employee in participants:
            notify(
                title=title,
                message=f'You are invited for a training commencing on {start_date}. {description}',
                users=employee,
                action_url=""
            )
            send_mail(subject=title, message=description, from_email=settings.EMAIL_HOST_USER,
                      recipient_list=[employee.email], fail_silently=False)

        messages.success(self.request, f"Training ({form.cleaned_data['title']}) updated successfully.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy("training_list")


@login_required(login_url='login')
def trash_training(request, pk, slug):
    training = get_object_or_404(Training, pk=pk, slug=slug)

    if not (is_supervisor(user=request.user) or is_hr(request.user)):
        messages.error(request, 'Access Denied.')
        return redirect('logout')

    if request.method == 'POST':
        messages.success(request, f"Training with ({training.title}) deleted successfully.")
        training.delete()
        return redirect('training_list')
    return render(request, 'hr/delete_confirmation.html', {'delete': training,
                                                           "cancel_url": reverse('training_list')})


@login_required(login_url='login')
def forum_list(request):
    threads_qs = ForumThread.objects.all().order_by("-created_at")
    paginator = Paginator(threads_qs, 10)  # 5 threads per page
    page_number = request.GET.get("page")
    threads = paginator.get_page(page_number)
    return render(request, "hr/forum_list.html", {"threads": threads})


@login_required(login_url='login')
def forum_detail(request, thread_id):
    thread = get_object_or_404(ForumThread, id=thread_id)
    posts = thread.posts.all().order_by("created_at")

    if request.method == "POST":
        form = ForumPostForm(request.POST)
        if form.is_valid():
            post = form.save(commit=False)
            post.thread = thread
            post.author = request.user
            post.save()
            return redirect("forum_detail", thread_id=thread.id)
    else:
        form = ForumPostForm()

    return render(request, "hr/forum_detail.html", {"thread": thread, "posts": posts, "form": form})


@login_required(login_url='login')
def new_thread(request):
    if request.method == "POST":
        form = ForumThreadForm(request.POST)
        if form.is_valid():
            thread = form.save(commit=False)
            thread.author = request.user
            thread.save()
            return redirect("forum_list")
    else:
        form = ForumThreadForm()
    return render(request, "hr/new_thread.html", {"form": form})


@login_required(login_url='login')
def delete_thread(request, thread_id):
    if request.user.is_superuser and is_hr(request.user):
        thread = get_object_or_404(ForumThread, id=thread_id)
    else:
        thread = get_object_or_404(ForumThread, id=thread_id, author=request.user)
    if request.method == "POST":
        thread.delete()
        return redirect("forum_list")
    return render(request, 'hr/delete_confirmation.html', {'delete': thread,
                                                           "cancel_url": reverse('forum_list')})


@login_required(login_url='login')
def delete_post(request, post_id):
    if request.user.is_superuser and is_hr(request.user):
        post = get_object_or_404(ForumPost, id=post_id)
    else:
        post = get_object_or_404(ForumPost, id=post_id, author=request.user)
    thread_id = post.thread.id
    if request.method == "POST":
        post.delete()
        return redirect("forum_detail", thread_id=thread_id)
    return render(request, 'hr/delete_confirmation.html', {'delete': post,
                                                           "cancel_url": reverse('forum_detail', args=[thread_id])})


@login_required(login_url='login')
def sitrep_list(request):
    sitrep_qs = SituationReport.objects.all().order_by("-created_at")
    paginator = Paginator(sitrep_qs, 5)  # 5 per page
    page_number = request.GET.get("page")
    reports = paginator.get_page(page_number)
    return render(request, "sitrep/sitrep_list.html", {"reports": reports})


@login_required(login_url='login')
def create_sitrep(request):
    if request.method == "POST":
        form = SituationReportForm(request.POST, request.FILES)
        if form.is_valid():
            sitrep = form.save(commit=False)
            sitrep.reporter = request.user
            sitrep.save()

            # Supervisors to notify
            supervisors = Profile.objects.filter(is_staff=True, is_superuser=True).values_list("email", flat=True)
            if supervisors:
                context = {
                    "reporter": request.user.get_full_name(),
                    "title": sitrep.title,
                    "description": sitrep.description,
                    "status": sitrep.status,
                    "url": request.build_absolute_uri(f"/sitrep/{sitrep.id}/"),
                    "year": datetime.now().year,
                }
                html_content = render_to_string("sitrep/email.html", context)
                text_content = strip_tags(html_content)

                email = EmailMultiAlternatives(
                    subject=f"🚨 New Situation Report: {sitrep.title}",
                    body=text_content,
                    from_email=f"{settings.EMAIL_HOST_USER}",
                    to=[f"{settings.EMAIL_HOST_USER}"],
                )
                email.attach_alternative(html_content, "text/html")
                email.send()

            # ✅ Create in-app notification
            notify(
                title=f"Situation Report '{sitrep.title}' submitted",
                message=f"Your Situation Report '{sitrep.title}' has been submitted.",
                users=request.user,
                action_url=""
            )

            messages.success(request, "✅ Situation Report submitted successfully.")
            return redirect("sitrep_list")
    else:
        form = SituationReportForm()
    return render(request, "sitrep/create_sitrep.html", {"form": form})


@login_required(login_url='login')
def sitrep_detail(request, pk):
    report = get_object_or_404(SituationReport, pk=pk)
    # Admin/supervisor status update
    if is_supervisor(request.user) and request.method == "POST":
        form = UpdateSitrepStatusForm(request.POST, instance=report)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Report status updated.")
            return redirect("sitrep_detail", pk=pk)
    else:
        form = UpdateSitrepStatusForm(instance=report) if is_supervisor(request.user) else None

    return render(request, "sitrep/sitrep_detail.html", {"report": report, "form": form})
